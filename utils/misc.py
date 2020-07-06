import functools
import os
import random
import warnings
from collections import OrderedDict
from datetime import datetime

import numpy as np
import torch
import torch.backends.cudnn as torchcudnn
from openpyxl import load_workbook, Workbook
from thop import profile
from torch.autograd.variable import Variable


class AvgMeter(object):
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count


def check_mkdir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def init_seed(seed: int):
    os.environ["PYTHONHASHSEED"] = str(seed)
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)  # if you are using multi-GPU.


def init_cudnn(benchmark=True):
    torchcudnn.enabled = True
    if benchmark:
        construct_print("We don't use the multi-training, so we will use the `cudnn.benchmark`")
    torchcudnn.benchmark = benchmark
    torchcudnn.deterministic = True
    construct_print(
        "You have chosen to seed training. "
        "This will turn on the CUDNN deterministic setting, "
        "which can slow down your training considerably! "
        "You may see unexpected behavior when restarting "
        "from checkpoints."
    )


def calc_flops(model, input_size):
    # USE_GPU = torch.cuda.is_available()
    USE_GPU = False

    def conv_hook(self, input, output):
        batch_size, input_channels, input_height, input_width = input[0].size()
        output_channels, output_height, output_width = output[0].size()

        kernel_ops = (
            self.kernel_size[0] * self.kernel_size[1] * (self.in_channels / self.groups) * (2 if multiply_adds else 1)
        )
        bias_ops = 1 if self.bias is not None else 0

        params = output_channels * (kernel_ops + bias_ops)
        flops = batch_size * params * output_height * output_width

        list_conv.append(flops)

    def linear_hook(self, input, output):
        batch_size = input[0].size(0) if input[0].dim() == 2 else 1

        weight_ops = self.weight.nelement() * (2 if multiply_adds else 1)
        bias_ops = self.bias.nelement()

        flops = batch_size * (weight_ops + bias_ops)
        list_linear.append(flops)

    def bn_hook(self, input, output):
        list_bn.append(input[0].nelement())

    def relu_hook(self, input, output):
        list_relu.append(input[0].nelement())

    def pooling_hook(self, input, output):
        batch_size, input_channels, input_height, input_width = input[0].size()
        output_channels, output_height, output_width = output[0].size()

        kernel_ops = self.kernel_size * self.kernel_size
        bias_ops = 0
        params = output_channels * (kernel_ops + bias_ops)
        flops = batch_size * params * output_height * output_width

        list_pooling.append(flops)

    def foo(net):
        childrens = list(net.children())
        if not childrens:
            if isinstance(net, torch.nn.Conv2d):
                net.register_forward_hook(conv_hook)
            if isinstance(net, torch.nn.Linear):
                net.register_forward_hook(linear_hook)
            if isinstance(net, torch.nn.BatchNorm2d):
                net.register_forward_hook(bn_hook)
            if isinstance(net, torch.nn.ReLU):
                net.register_forward_hook(relu_hook)
            if isinstance(net, torch.nn.MaxPool2d) or isinstance(net, torch.nn.AvgPool2d):
                net.register_forward_hook(pooling_hook)
            return
        for c in childrens:
            foo(c)

    multiply_adds = False
    list_conv, list_bn, list_relu, list_linear, list_pooling = [], [], [], [], []
    foo(model)

    if "0.4." in torch.__version__ or "1.0" in torch.__version__:
        if USE_GPU:
            input = torch.cuda.FloatTensor(torch.rand(2, 3, input_size, input_size).cuda())
        else:
            input = torch.FloatTensor(torch.rand(2, 3, input_size, input_size))
    else:
        input = Variable(torch.rand(2, 3, input_size, input_size), requires_grad=True)
    _ = model(input)

    total_flops = sum(list_conv) + sum(list_linear) + sum(list_bn) + sum(list_relu) + sum(list_pooling)

    print("  + Number of FLOPs: %.2fG" % (total_flops / 1e9 / 2))


def count_params(model, input_size=224):
    # param_sum = 0
    # with open('models.txt', 'w') as fm:
    #     fm.write(str(model))
    calc_flops(model, input_size)

    model_parameters = filter(lambda p: p.requires_grad, model.parameters())
    params = sum([np.prod(p.size()) for p in model_parameters])

    print("The network has {} params.".format(params))


def print_head_and_tail(local_rank):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            if local_rank == 0:
                construct_print(f"{datetime.now()}: Start...")
            end_str = func(*args, **kwargs)
            if local_rank == 0:
                construct_print(f"{datetime.now()}: {end_str}...")

        return wrapper

    return decorator


def construct_exp_name(arg_dict: dict):
    # bs_16_lr_0.05_e30_noamp_2gpu_noms_352
    focus_item = OrderedDict(
        {
            "input_size": "size",
            "batch_size": "bs",
            "lr": "lr",
            "epoch_num": "e",
            "use_amp": "amp",
            "is_distributed": "dist",
            "size_list": "ms",
            "version": "v",
        }
    )
    exp_name = f"{arg_dict['model']}"
    for k, v in focus_item.items():
        item = arg_dict[k]
        if isinstance(item, bool):
            item = "Y" if item else "N"
        elif isinstance(item, (list, tuple)):
            item = "Y" if item else "N"  # 只是判断是否飞空
        elif isinstance(item, str):
            if not item:
                continue
        elif item == None:
            item = "N"

        if isinstance(item, str):
            item = item.lower()

        exp_name += f"_{v.upper()}{item}"
    return exp_name


def construct_path_dict(proj_root, exp_name, xlsx_name):
    ckpt_path = os.path.join(proj_root, "output")

    pth_log_path = os.path.join(ckpt_path, exp_name)
    tb_path = os.path.join(pth_log_path, "tb")
    save_path = os.path.join(pth_log_path, "pre")
    pth_path = os.path.join(pth_log_path, "pth")

    final_full_model_path = os.path.join(pth_path, "checkpoint_final.pth.tar")
    final_state_path = os.path.join(pth_path, "state_final.pth")

    tr_log_path = os.path.join(pth_log_path, f"tr_{str(datetime.now())[:10]}.txt")
    te_log_path = os.path.join(pth_log_path, f"te_{str(datetime.now())[:10]}.txt")
    cfg_log_path = os.path.join(pth_log_path, f"cfg_{str(datetime.now())[:10]}.txt")
    trainer_log_path = os.path.join(pth_log_path, f"trainer_{str(datetime.now())[:10]}.txt")
    xlsx_path = os.path.join(ckpt_path, xlsx_name)

    path_config = {
        "ckpt_path": ckpt_path,
        "pth_log": pth_log_path,
        "tb": tb_path,
        "save": save_path,
        "pth": pth_path,
        "final_full_net": final_full_model_path,
        "final_state_net": final_state_path,
        "tr_log": tr_log_path,
        "te_log": te_log_path,
        "cfg_log": cfg_log_path,
        "trainer_log": trainer_log_path,
        "xlsx": xlsx_path,
    }
    return path_config


def get_FLOPs_Params(model, channels, input_size=320, mode="print"):
    input = torch.randn(1, channels, input_size, input_size).cuda()
    flops, params = profile(model, inputs=(input,))

    if mode == "print":
        print(f"    + Number of FLOPs: {flops / 1e9:.2f}G\n  The network has {params} params.")
    elif mode == "return":
        return flops, params
    elif mode == "print&return":
        msg = f"    + Number of FLOPs: {flops / 1e9:.2f}G\n  The network has {params} params."
        print(msg)
        return msg
    else:
        raise NotImplementedError


def make_xlsx(xlsx_path):
    num_metrics = len(metric_list)
    num_datasets = len(dataset_list)

    # 创建一个Workbook对象
    wb = Workbook()
    # 创建一个Sheet对象
    sheet = wb.create_sheet(title="Results", index=0)
    # 获取活动的sheet
    sheet["A1"] = "name_dataset"
    sheet["A2"] = "num_dataset"

    for i, dataset_name in enumerate(dataset_list):
        if (i * num_metrics + 1) // 26 == 0:
            start_region_idx = f"{chr(ord('A') + (i * num_metrics + 1) % 26)}1"
        else:
            start_region_idx = (
                f"{chr(ord('A') + (i * num_metrics + 1) // 26 - 1)}" f"{chr(ord('A') + (i * num_metrics + 1) % 26)}1"
            )
        if ((i + 1) * num_metrics) // 26 == 0:
            end_region_idx = f"{chr(ord('A') + ((i + 1) * num_metrics) % 26)}1"
        else:
            end_region_idx = (
                f"{chr(ord('A') + ((i + 1) * num_metrics) // 26 - 1)}"
                f"{chr(ord('A') + ((i + 1) * num_metrics) % 26)}1"
            )
        region_idx = f"{start_region_idx}:{end_region_idx}"
        sheet.merge_cells(region_idx)  # 合并一行中的几个单元格
        sheet[start_region_idx] = dataset_name

        # 构造第二行数据
        start_region_idx = start_region_idx.replace("1", "2")
        sheet[start_region_idx] = dataset_num_list[i]

    # 构造第三行数据
    third_row = ["metrics"] + metric_list * num_datasets
    sheet.append(third_row)

    # 最后保存workbook
    wb.save(xlsx_path)


def write_xlsx(model_name, data):
    """
    向xlsx文件中写入数据

    :param model_name: 模型名字
    :param data: 数据信息，包含数据集名字和对应的测试结果
    """

    num_metrics = len(metric_list)
    num_datasets = len(dataset_list)

    # 必须先得由前面的部分进行xlsx文件的创建，确保前三行OK满足要求，后面的操作都是从第四行开始的
    wb = load_workbook(path_config["xlsx"])
    assert "Results" in wb.sheetnames, "Please make sure you are " "working with xlsx files " "created by `make_xlsx`"
    sheet = wb["Results"]
    num_cols = num_metrics * num_datasets + 1

    if model_name in sheet["A"]:
        # 说明，该模型已经存在条目中，只需要更新对应的数据集结果即可
        idx_insert_row = sheet["A"].find(model_name)
    else:
        idx_insert_row = len(sheet["A"]) + 1
        sheet.cell(row=idx_insert_row, column=1, value=model_name)

    for dataset_name in data.keys():
        # 遍历每个单元格
        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=num_cols, max_row=1):
            for cell in row:
                if cell.value == dataset_name:
                    for i in range(num_metrics):
                        matric_name = sheet.cell(row=3, column=cell.column + i).value
                        sheet.cell(
                            row=idx_insert_row, column=cell.column + i, value=data[dataset_name][matric_name],
                        )
    wb.save(path_config["xlsx"])


def construct_print(out_str: str, total_length: int = 80):
    if len(out_str) >= total_length:
        extended_str = "=="
    else:
        extended_str = "=" * ((total_length - len(out_str)) // 2 - 4)
    out_str = f" {extended_str}>> {out_str} <<{extended_str} "
    print(out_str)


def write_data_to_file(data_str, file_path):
    with open(file_path, encoding="utf-8", mode="a") as f:
        f.write(data_str + "\n")


if __name__ == "__main__":
    print("=" * 8)
    out_str = "lartpang"
    construct_print(out_str, total_length=8)
